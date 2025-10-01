#!/usr/bin/env python3
"""
Excel Template Generator for Manual Data Entry
Creates an Excel file with data validation for easy manual entry of 20 client plans
"""

import pandas as pd
from pathlib import Path
from datetime import datetime


def create_excel_template(output_path: str = 'data/plan_data_template.xlsx'):
    """
    Create Excel template with headers, descriptions, and example row.

    Args:
        output_path: Path where Excel template will be saved
    """

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    print("Creating Excel template for manual data entry...")

    # Define column headers with descriptions
    columns = [
        'client_id',
        'client_name',
        'industry',
        'employee_count',
        'state',
        'eligibility',
        'match_formula',
        'match_effective_rate',
        'match_eligibility_criteria',
        'match_last_day_work_rule',
        'match_true_up',
        'match_contribution_frequency',
        'nonelective_formula',
        'nonelective_eligibility_criteria',
        'nonelective_last_day_work_rule',
        'nonelective_contribution_frequency',
        'auto_enrollment_enabled',
        'auto_enrollment_rate',
        'auto_enrollment_effective_year',
        'auto_escalation_enabled',
        'auto_escalation_cap',
        'vesting_schedule',
        'data_source',
        'notes'
    ]

    # Column descriptions (will be in row 2)
    descriptions = [
        'Unique client ID (e.g., CLIENT-001)',
        'Company name',
        'healthcare, higher_ed, manufacturing, or other',
        'Number of employees',
        'Two-letter state code (CA, NY, etc.)',
        'e.g., "Immediate", "1 year", "1000 hours"',
        'e.g., "100% up to 3%" or "50% up to 6%"',
        'Effective match rate as % (e.g., 3.0)',
        'Eligibility criteria for match',
        'Yes or No - Last day work rule for match',
        'Yes or No - Match true-up enabled',
        'e.g., "per pay period", "monthly", "annually"',
        'Non-elective contribution formula',
        'Eligibility criteria for non-elective',
        'Yes or No - Last day work rule for non-elective',
        'e.g., "per pay period", "monthly", "annually"',
        'Yes or No',
        'Default rate as % (e.g., 3.0) - leave blank if No AE',
        'Year AE became effective (e.g., 1900 for all, 2024 for new hires)',
        'Yes or No - Auto-escalation enabled',
        'Auto-escalation cap as % (e.g., 10.0)',
        'e.g., "Immediate", "3-year cliff", "2-6 graded"',
        'ManualEntry (default)',
        'Any notes about data quality or source'
    ]

    # Example data (row 3)
    example_row = [
        'CLIENT-001',
        'ABC University',
        'higher_ed',
        2500,
        'CA',
        'Immediate',
        '100% up to 4%',
        4.0,
        'All eligible employees',
        'No',
        'Yes',
        'per pay period',
        '3% of compensation',
        'All eligible employees',
        'No',
        'annually',
        'Yes',
        3.0,
        1900,
        'Yes',
        10.0,
        'Immediate',
        'ManualEntry',
        'Data from 2023 Form 5500'
    ]

    # Create DataFrame with headers, descriptions, and example
    df = pd.DataFrame([descriptions, example_row], columns=columns)

    # Create Excel writer with openpyxl engine for formatting
    from openpyxl.styles import Font, PatternFill

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Plan Data', index=False, startrow=1)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Plan Data']

        # Write column headers in row 1
        for col_num, column_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Format description row (row 2)
        for col_num in range(1, len(columns) + 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.font = Font(italic=True, size=9, color='666666')

        # Format example row (row 3)
        for col_num in range(1, len(columns) + 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')

        # Add data validation for dropdowns
        from openpyxl.worksheet.datavalidation import DataValidation

        # Industry dropdown (column C)
        industry_validation = DataValidation(
            type="list",
            formula1='"healthcare,higher_ed,manufacturing,other"',
            allow_blank=False
        )
        worksheet.add_data_validation(industry_validation)
        industry_validation.add(f'C4:C50')

        # Yes/No dropdowns
        yes_no_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        worksheet.add_data_validation(yes_no_validation)
        # Apply to: match_last_day_work_rule (J), match_true_up (K),
        # nonelective_last_day_work_rule (O), auto_enrollment_enabled (Q),
        # auto_escalation_enabled (T)
        yes_no_validation.add(f'J4:J50')
        yes_no_validation.add(f'K4:K50')
        yes_no_validation.add(f'O4:O50')
        yes_no_validation.add(f'Q4:Q50')
        yes_no_validation.add(f'T4:T50')

        # Set column widths
        column_widths = {
            'A': 15,  # client_id
            'B': 25,  # client_name
            'C': 18,  # industry
            'D': 15,  # employee_count
            'E': 8,   # state
            'F': 20,  # eligibility
            'G': 20,  # match_formula
            'H': 20,  # match_effective_rate
            'I': 25,  # match_eligibility_criteria
            'J': 22,  # match_last_day_work_rule
            'K': 15,  # match_true_up
            'L': 25,  # match_contribution_frequency
            'M': 25,  # nonelective_formula
            'N': 30,  # nonelective_eligibility_criteria
            'O': 28,  # nonelective_last_day_work_rule
            'P': 30,  # nonelective_contribution_frequency
            'Q': 22,  # auto_enrollment_enabled
            'R': 22,  # auto_enrollment_rate
            'S': 28,  # auto_enrollment_effective_year
            'T': 22,  # auto_escalation_enabled
            'U': 20,  # auto_escalation_cap
            'V': 25,  # vesting_schedule
            'W': 15,  # data_source
            'X': 30   # notes
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Freeze top 3 rows (headers, descriptions, example)
        worksheet.freeze_panes = 'A4'

        # Add instructions sheet
        instructions_data = {
            'Step': [1, 2, 3, 4, 5, 6, 7],
            'Instructions': [
                'Start entering data in Row 4 (below the example row)',
                'Use the dropdown menus for industry and Yes/No fields',
                'For percentage fields: Enter the % value (e.g., 3.0 for 3%, 4.5 for 4.5%)',
                'For auto_enrollment_effective_year: Enter 1900 if applies to all employees, or year like 2024 for new hires after that date',
                'Leave auto_enrollment_rate blank if auto_enrollment_enabled = No',
                'Fill in all 20 clients (rows 4-23)',
                'Save the file and run: python excel_to_duckdb.py'
            ]
        }

        instructions_df = pd.DataFrame(instructions_data)
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

        # Format instructions sheet
        inst_worksheet = writer.sheets['Instructions']
        inst_worksheet.column_dimensions['A'].width = 8
        inst_worksheet.column_dimensions['B'].width = 80

        for row in inst_worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # Add field reference sheet
        field_reference = {
            'Field Name': [
                'client_id',
                'client_name',
                'industry',
                'employee_count',
                'state',
                'eligibility',
                'match_formula',
                'match_effective_rate',
                'match_eligibility_criteria',
                'match_last_day_work_rule',
                'match_true_up',
                'match_contribution_frequency',
                'nonelective_formula',
                'nonelective_eligibility_criteria',
                'nonelective_last_day_work_rule',
                'nonelective_contribution_frequency',
                'auto_enrollment_enabled',
                'auto_enrollment_rate',
                'auto_enrollment_effective_year',
                'auto_escalation_enabled',
                'auto_escalation_cap',
                'vesting_schedule',
                'data_source',
                'notes'
            ],
            'Data Type': [
                'Text',
                'Text',
                'Dropdown',
                'Number',
                'Text (2 letters)',
                'Text',
                'Text',
                'Number (decimal)',
                'Text',
                'Dropdown',
                'Dropdown',
                'Text',
                'Text',
                'Text',
                'Dropdown',
                'Text',
                'Dropdown',
                'Number (decimal)',
                'Number (year)',
                'Dropdown',
                'Number (decimal)',
                'Text',
                'Text',
                'Text'
            ],
            'Required': [
                'Yes',
                'Yes',
                'Yes',
                'Yes',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'No',
                'Yes',
                'Conditional',
                'Conditional',
                'No',
                'Conditional',
                'No',
                'Yes',
                'No'
            ],
            'Notes': [
                'Must be unique (e.g., CLIENT-001, CLIENT-002)',
                'Company legal name',
                'Options: healthcare, higher_ed, manufacturing, other',
                'Total employee count (not just participants)',
                'Two-letter state code (CA, NY, TX, etc.) or leave blank',
                'Eligibility for plan entry (e.g., "Immediate", "1 year", "1000 hours")',
                'Human-readable format (e.g., "100% up to 3%", "50% up to 6%")',
                'Effective match as % (3.0 = 3%, 4.5 = 4.5%)',
                'Eligibility criteria for employer match',
                'Options: Yes, No. Last day work rule requirement for match',
                'Options: Yes, No. Match true-up available',
                'Frequency of match contributions (e.g., "per pay period", "monthly", "annually")',
                'Non-elective contribution formula if applicable',
                'Eligibility criteria for non-elective contribution',
                'Options: Yes, No. Last day work rule requirement for non-elective',
                'Frequency of non-elective contributions',
                'Options: Yes, No',
                'Required if auto_enrollment_enabled = Yes. Enter as % (3.0 = 3%)',
                'Year auto-enrollment became effective. 1900 = applies to all employees',
                'Options: Yes, No. Auto-escalation feature enabled',
                'Required if auto_escalation_enabled = Yes. Cap as % (10.0 = 10%)',
                'Vesting schedule (e.g., "Immediate", "3-year cliff", "2-6 graded")',
                'Default: ManualEntry',
                'Any context about data quality or extraction challenges'
            ]
        }

        field_ref_df = pd.DataFrame(field_reference)
        field_ref_df.to_excel(writer, sheet_name='Field Reference', index=False)

        # Format field reference sheet
        ref_worksheet = writer.sheets['Field Reference']
        ref_worksheet.column_dimensions['A'].width = 25
        ref_worksheet.column_dimensions['B'].width = 20
        ref_worksheet.column_dimensions['C'].width = 12
        ref_worksheet.column_dimensions['D'].width = 60

        for row in ref_worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    print(f"✓ Excel template created: {output_file.absolute()}")
    print(f"\n{'='*60}")
    print("Template includes:")
    print("  • Sheet 1: Plan Data (with example row and data validation)")
    print("  • Sheet 2: Instructions (step-by-step guide)")
    print("  • Sheet 3: Field Reference (detailed field descriptions)")
    print(f"{'='*60}")
    print("\nNext steps:")
    print("  1. Open the Excel file")
    print("  2. Start entering data in Row 4")
    print("  3. Fill in 20 client plans (rows 4-23)")
    print("  4. Save and run: python excel_to_duckdb.py")

    return output_path


if __name__ == "__main__":
    print("PlanWise Design Matrix - Excel Template Generator")
    print("=" * 60)

    template_path = create_excel_template()

    print(f"\n✓ Ready to start manual data entry!")
    print(f"✓ Template location: {Path(template_path).absolute()}")